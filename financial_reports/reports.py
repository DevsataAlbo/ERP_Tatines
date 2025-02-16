# financial_reports/reports.py
import io
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from django.db.models import Sum, F, Q, Count, Avg, Case, When
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
from sales.models import Sale, SaleDetail
from expenses.models import Expense

class ReportGenerator:
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=16,
            spaceAfter=30
        )

    def _clean_data(self, value):
        """Limpia valores numéricos para reportes"""
        if value is None:
            return 0
        if isinstance(value, Decimal):
            return float(value)
        return value

    # Agregar también los métodos faltantes al get_report_data
    def get_report_data(self, report_type, date_from, date_to):
        """Obtiene los datos según el tipo de reporte"""
        if report_type == 'income_statement':
            return self._get_income_statement_data(date_from, date_to)
        elif report_type == 'sales_detail':
            return self._get_sales_detail_data(date_from, date_to)
        elif report_type == 'expense_detail':
            return self._get_expense_detail_data(date_from, date_to)
        elif report_type == 'profitability':
            return self._get_profitability_data(date_from, date_to)
        elif report_type == 'sales_kpi':
            return self._get_sales_kpi_data(date_from, date_to)
        elif report_type == 'expense_kpi':
            return self._get_expense_kpi_data(date_from, date_to)
        elif report_type == 'profit_kpi':
            return self._get_profit_kpi_data(date_from, date_to)
        elif report_type == 'sales_comparison':
            return self._get_sales_comparison_data(date_from, date_to)
        elif report_type == 'expense_comparison':
            return self._get_expense_comparison_data(date_from, date_to)
        elif report_type == 'profit_comparison':
            return self._get_profit_comparison_data(date_from, date_to)
        else:
            raise ValueError(f"Tipo de reporte no válido: {report_type}")

    def generate_pdf(self, data, report_type, date_from=None, date_to=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Título y fecha
        title = f"Reporte de {report_type.replace('_', ' ').title()}"
        date_range = f"Período: {date_from} - {date_to}" if date_from and date_to else ""
        
        elements.append(Paragraph(title, self.title_style))
        elements.append(Paragraph(date_range, self.styles['Normal']))
        elements.append(Spacer(1, 20))

        # Agregar datos según tipo de reporte
        if report_type == 'income_statement':
            elements.extend(self._generate_income_statement_pdf(data))
        elif report_type == 'sales_detail':
            elements.extend(self._generate_sales_detail_pdf(data))
        elif report_type == 'expense_detail':
            elements.extend(self._generate_expense_detail_pdf(data))
        elif report_type == 'profitability':
            elements.extend(self._generate_profitability_pdf(data))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_excel(self, data, report_type, date_from=None, date_to=None):
        wb = Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()

        # Estilo para encabezados
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        header_font = Font(bold=True)

        # Generar hojas según tipo de reporte
        if report_type == 'income_statement':
            self._generate_income_statement_excel(ws, data, header_fill, header_font)
        elif report_type == 'sales_detail':
            self._generate_sales_detail_excel(ws, data, header_fill, header_font)
        elif report_type == 'expense_detail':
            self._generate_expense_detail_excel(ws, data, header_fill, header_font)
        elif report_type == 'profitability':
            self._generate_profitability_excel(ws, data, header_fill, header_font)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def _get_income_statement_data(self, date_from, date_to):
        """Datos para Estado de Resultados"""
        sales = Sale.objects.filter(status='COMPLETED')
        expenses = Expense.objects.all()
        
        if date_from:
            sales = sales.filter(date__gte=date_from)
            expenses = expenses.filter(date__gte=date_from)
        if date_to:
            sales = sales.filter(date__lte=date_to)
            expenses = expenses.filter(date__lte=date_to)

        # Calcular total y neto
        total = self._clean_data(sales.aggregate(total=Sum('total'))['total'])
        neto = total / 1.19  # Calculamos el neto dividiendo por 1.19 (IVA 19%)
        iva = total - neto

        sales_data = {
            'total': total,
            'net': neto,
            'tax': iva,
            'by_method': list(sales.values('payment_method').annotate(
                total=Sum('total'),
                count=Count('id')
            ))
        }

        costs_data = {
            'products': self._clean_data(sales.aggregate(
                total=Sum(F('saledetail__quantity') * F('saledetail__product__purchase_price'))
            )['total']),
            'commissions': self._clean_data(sales.aggregate(
                total=Sum('commission_amount')
            )['total'])
        }

        expenses_data = {
            'total': self._clean_data(expenses.aggregate(total=Sum('amount'))['total']),
            'by_category': list(expenses.values('category__name').annotate(
                total=Sum('amount'),
                count=Count('id')
            ))
        }

        return {
            'sales': sales_data,
            'costs': costs_data,
            'expenses': expenses_data
        }

    def _get_sales_detail_data(self, date_from, date_to):
        """Datos para Detalle de Ventas"""
        sales = Sale.objects.filter(status='COMPLETED')
        if date_from:
            sales = sales.filter(date__gte=date_from)
        if date_to:
            sales = sales.filter(date__lte=date_to)

        # Para calcular el profit
        sale_details = SaleDetail.objects.filter(
            sale__in=sales
        ).values('product__name').annotate(
            total_quantity=Sum('quantity'),
            total_amount=Sum(F('quantity') * F('unit_price')),
            total_cost=Sum(F('quantity') * F('product__purchase_price'))
        )

        # Calcular profit manualmente
        for detail in sale_details:
            detail['profit'] = detail['total_amount'] - detail['total_cost']

        # Resumen diario modificado
        daily_summary = sales.annotate(
            day=TruncDay('date')
        ).values('day').annotate(
            total_sales=Sum('total'),
            transaction_count=Count('id')
        ).order_by('day')

        # Calcular ticket promedio manualmente
        for day in daily_summary:
            day['avg_ticket'] = day['total_sales'] / day['transaction_count'] if day['transaction_count'] > 0 else 0

        return {
            'sales_list': list(sales.annotate(
                products_count=Count('saledetail'),
                items_count=Sum('saledetail__quantity')
            ).order_by('-date')),
            'by_product': list(sale_details.order_by('-total_amount')[:10]),
            'daily_summary': list(daily_summary)
        }
    
    def _get_expense_detail_data(self, date_from, date_to):
        """Datos para Detalle de Gastos"""
        expenses = Expense.objects.all()
        if date_from:
            expenses = expenses.filter(date__gte=date_from)
        if date_to:
            expenses = expenses.filter(date__lte=date_to)
        return {
            'expense_list': list(expenses.select_related('category').order_by('-date')),
            'by_category': list(expenses.values('category__name').annotate(
                total=Sum('amount'),
                count=Count('id'),
                avg=Avg('amount')
            ).order_by('-total')),
            'monthly_summary': list(expenses.annotate(
                month=TruncMonth('date')
            ).values('month').annotate(
                total=Sum('amount'),
                count=Count('id')
            ).order_by('month'))
        }
    def _get_profitability_data(self, date_from, date_to):
        """Datos para Análisis de Rentabilidad"""
        sales = Sale.objects.filter(status='COMPLETED')
        if date_from:
            sales = sales.filter(date__gte=date_from)
        if date_to:
            sales = sales.filter(date__lte=date_to)
        by_product = list(SaleDetail.objects.filter(
            sale__in=sales
        ).values('product__name').annotate(
            total_sales=Sum(F('quantity') * F('unit_price')),
            total_cost=Sum(F('quantity') * F('product__purchase_price')),
            profit=F('total_sales') - F('total_cost'),
            margin=Case(
                When(total_sales=0, then=0),
                default=F('profit') * 100.0 / F('total_sales')
            )
        ).order_by('-profit'))
        by_category = list(SaleDetail.objects.filter(
            sale__in=sales
        ).values('product__category__name').annotate(
            total_sales=Sum(F('quantity') * F('unit_price')),
            total_cost=Sum(F('quantity') * F('product__purchase_price')),
            profit=F('total_sales') - F('total_cost')
        ).order_by('-profit'))
        return {
            'by_product': [
                {**item, 
                'total_sales': self._clean_data(item['total_sales']),
                'total_cost': self._clean_data(item['total_cost']),
                'profit': self._clean_data(item['profit']),
                'margin': self._clean_data(item['margin'])}
                for item in by_product
            ],
            'by_category': [
                {**item, 
                'total_sales': self._clean_data(item['total_sales']),
                'total_cost': self._clean_data(item['total_cost']),
                'profit': self._clean_data(item['profit'])}
                for item in by_category
            ]
        }

    def _generate_income_statement_pdf(self, data):
            elements = []
            
            # Ventas
            elements.append(Paragraph('Resumen de Ventas', self.styles['Heading2']))
            elements.append(Spacer(1, 12))
            
            sales_data = [
                ['', 'Monto'],
                ['Ventas Totales', f"$ {data['sales']['total']:,.0f}"],
                ['Ventas Netas', f"$ {data['sales']['net']:,.0f}"],
                ['IVA', f"$ {data['sales']['tax']:,.0f}"]
            ]
            
            sales_table = Table(sales_data, colWidths=[300, 200])
            sales_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('BOX', (0, 0), (-1, -1), 2, colors.black)
            ]))
            elements.append(sales_table)
            elements.append(Spacer(1, 20))

            # Costos y Gastos
            elements.append(Paragraph('Costos y Gastos', self.styles['Heading2']))
            elements.append(Spacer(1, 12))
            
            costs_data = [
                ['Concepto', 'Monto'],
                ['Costo de Productos', f"$ {data['costs']['products']:,.0f}"],
                ['Comisiones', f"$ {data['costs']['commissions']:,.0f}"],
                ['Gastos Operacionales', f"$ {data['expenses']['total']:,.0f}"]
            ]
            
            costs_table = Table(costs_data, colWidths=[300, 200])
            costs_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(costs_table)

            return elements

    def _generate_sales_detail_excel(self, ws, data, header_fill, header_font):
            ws['A1'] = 'Detalle de Ventas'
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:D1')
            
            # Top 10 productos
            current_row = 3
            headers = ['Producto', 'Cantidad', 'Total', 'Ganancia']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            for product in data['by_product']:
                current_row += 1
                ws.cell(row=current_row, column=1, value=product['product__name'])
                ws.cell(row=current_row, column=2, value=self._clean_data(product['total_quantity']))
                ws.cell(row=current_row, column=3, value=self._clean_data(product['total_amount']))
                ws.cell(row=current_row, column=4, value=self._clean_data(product['profit']))

            # Formato números
            for row in range(4, current_row + 1):
                for col in [2, 3, 4]:
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = '#,##0'

    def _generate_expense_detail_excel(self, ws, data, header_fill, header_font):
        ws['A1'] = 'Detalle de Gastos'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')
        
        current_row = 3
        headers = ['Categoría', 'Total', 'Cantidad', 'Promedio']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for category in data['by_category']:
            current_row += 1
            ws.cell(row=current_row, column=1, value=category['category__name'])
            ws.cell(row=current_row, column=2, value=self._clean_data(category['total']))
            ws.cell(row=current_row, column=3, value=category['count'])
            ws.cell(row=current_row, column=4, value=self._clean_data(category['avg']))
        # Formato números
        for row in range(4, current_row + 1):
            for col in [2, 4]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
                
    def _generate_profitability_excel(self, ws, data, header_fill, header_font):
        ws['A1'] = 'Análisis de Rentabilidad'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        current_row = 3
        headers = ['Producto', 'Ventas', 'Costo', 'Ganancia', 'Margen']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        for product in data['by_product']:
            current_row += 1
            ws.cell(row=current_row, column=1, value=product['product__name'])
            ws.cell(row=current_row, column=2, value=product['total_sales'])
            ws.cell(row=current_row, column=3, value=product['total_cost'])
            ws.cell(row=current_row, column=4, value=product['profit'])
            ws.cell(row=current_row, column=5, value=product['margin'])
        # Formato números
        for row in range(4, current_row + 1):
            for col in [2, 3, 4]:
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0'
            cell = ws.cell(row=row, column=5)
            cell.number_format = '0.0%'

        # Agregar al final de la clase ReportGenerator

    def _get_sales_kpi_data(self, date_from, date_to):
        """KPIs de ventas"""
        sales = Sale.objects.filter(status='COMPLETED')
        if date_from:
            sales = sales.filter(date__gte=date_from)
        if date_to:
            sales = sales.filter(date__lte=date_to)

        # Calcular período anterior para comparación
        if date_from and date_to:
            days_diff = (date_to - date_from).days
            prev_date_to = date_from
            prev_date_from = date_from - timezone.timedelta(days=days_diff)
            prev_sales = Sale.objects.filter(
                status='COMPLETED',
                date__gte=prev_date_from,
                date__lte=prev_date_to
            )
        else:
            prev_sales = None

        current_total = sales.aggregate(total=Sum('total'))['total'] or 0
        prev_total = prev_sales.aggregate(total=Sum('total'))['total'] if prev_sales else 0
        
        return {
            'total_sales': self._clean_data(current_total),
            'average_ticket': self._clean_data(current_total / sales.count() if sales.count() > 0 else 0),
            'ticket_trend': self._calculate_trend(current_total, prev_total),
            'total_transactions': sales.count(),
            'daily_average': self._clean_data(current_total / len(set(sales.values_list('date', flat=True))) if sales.exists() else 0)
        }

    def _get_expense_kpi_data(self, date_from, date_to):
        """KPIs de gastos"""
        expenses = Expense.objects.all()
        if date_from:
            expenses = expenses.filter(date__gte=date_from)
        if date_to:
            expenses = expenses.filter(date__lte=date_to)

        total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
        
        return {
            'total_expenses': self._clean_data(total_expenses),
            'by_category': list(expenses.values('category__name').annotate(
                total=Sum('amount'),
                percentage=F('total') * 100.0 / total_expenses if total_expenses > 0 else 0
            ).order_by('-total')),
            'monthly_average': self._clean_data(total_expenses / 12),  # Promedio mensual
            'daily_average': self._clean_data(total_expenses / 30)  # Promedio diario aproximado
        }

    def _get_profit_kpi_data(self, date_from, date_to):
        """KPIs de rentabilidad"""
        sales_data = self._get_sales_kpi_data(date_from, date_to)
        expense_data = self._get_expense_kpi_data(date_from, date_to)
        
        total_revenue = sales_data['total_sales']
        total_expenses = expense_data['total_expenses']
        gross_profit = total_revenue - total_expenses
        
        return {
            'gross_profit': self._clean_data(gross_profit),
            'profit_margin': self._clean_data((gross_profit / total_revenue * 100) if total_revenue > 0 else 0),
            'roi': self._clean_data((gross_profit / total_expenses * 100) if total_expenses > 0 else 0),
            'break_even': self._clean_data(total_expenses * 1.2)  # Punto de equilibrio con 20% de margen
        }

    def _calculate_trend(self, current, previous):
        """Calcula la tendencia porcentual entre dos valores"""
        if not previous:
            return 0
        return ((current - previous) / previous * 100) if previous != 0 else 0

    # Métodos de comparación
    def _get_sales_comparison_data(self, date_from, date_to):
        """Comparación de ventas"""
        current_period = self._get_sales_detail_data(date_from, date_to)
        
        # Calcular período anterior
        if date_from and date_to:
            days_diff = (date_to - date_from).days
            prev_date_to = date_from
            prev_date_from = date_from - timezone.timedelta(days=days_diff)
            previous_period = self._get_sales_detail_data(prev_date_from, prev_date_to)
        else:
            previous_period = None
        
        return {
            'current_period': current_period,
            'previous_period': previous_period,
            'trends': {
                'sales': self._calculate_trend(
                    sum(day['total'] for day in current_period['daily_summary']),
                    sum(day['total'] for day in previous_period['daily_summary']) if previous_period else 0
                ),
                'transactions': self._calculate_trend(
                    len(current_period['sales_list']),
                    len(previous_period['sales_list']) if previous_period else 0
                )
            }
        }
        
    def _get_expense_comparison_data(self, date_from, date_to):
        """Comparación de gastos entre períodos"""
        current_period = self._get_expense_detail_data(date_from, date_to)
        
        # Calcular período anterior
        if date_from and date_to:
            days_diff = (date_to - date_from).days
            prev_date_to = date_from
            prev_date_from = date_from - timezone.timedelta(days=days_diff)
            previous_period = self._get_expense_detail_data(prev_date_from, prev_date_to)
        else:
            previous_period = None
        
        return {
            'current_period': current_period,
            'previous_period': previous_period,
            'trends': {
                'total_expenses': self._calculate_trend(
                    sum(exp['total'] for exp in current_period['by_category']),
                    sum(exp['total'] for exp in previous_period['by_category']) if previous_period else 0
                ),
                'by_category': [
                    {
                        'category': cat['category__name'],
                        'current': self._clean_data(cat['total']),
                        'previous': self._clean_data(
                            next(
                                (p['total'] for p in previous_period['by_category'] 
                                if p['category__name'] == cat['category__name']),
                                0
                            ) if previous_period else 0
                        ),
                        'trend': self._calculate_trend(
                            cat['total'],
                            next(
                                (p['total'] for p in previous_period['by_category'] 
                                if p['category__name'] == cat['category__name']),
                                0
                            ) if previous_period else 0
                        )
                    }
                    for cat in current_period['by_category']
                ]
            }
        }

    def _get_profit_comparison_data(self, date_from, date_to):
        """Comparación de rentabilidad entre períodos"""
        current_period = self._get_profitability_data(date_from, date_to)
        
        # Calcular período anterior
        if date_from and date_to:
            days_diff = (date_to - date_from).days
            prev_date_to = date_from
            prev_date_from = date_from - timezone.timedelta(days=days_diff)
            previous_period = self._get_profitability_data(prev_date_from, prev_date_to)
        else:
            previous_period = None

        # Calcular tendencias por producto
        products_comparison = []
        for curr_prod in current_period['by_product']:
            prev_prod = next(
                (p for p in previous_period['by_product'] 
                if p['product__name'] == curr_prod['product__name']),
                None
            ) if previous_period else None

            products_comparison.append({
                'product': curr_prod['product__name'],
                'current_profit': self._clean_data(curr_prod['profit']),
                'previous_profit': self._clean_data(prev_prod['profit'] if prev_prod else 0),
                'current_margin': self._clean_data(curr_prod['margin']),
                'previous_margin': self._clean_data(prev_prod['margin'] if prev_prod else 0),
                'profit_trend': self._calculate_trend(
                    curr_prod['profit'],
                    prev_prod['profit'] if prev_prod else 0
                ),
                'margin_trend': self._calculate_trend(
                    curr_prod['margin'],
                    prev_prod['margin'] if prev_prod else 0
                )
            })

        return {
            'current_period': current_period,
            'previous_period': previous_period,
            'products_comparison': products_comparison,
            'summary': {
                'total_profit_trend': self._calculate_trend(
                    sum(p['profit'] for p in current_period['by_product']),
                    sum(p['profit'] for p in previous_period['by_product']) if previous_period else 0
                ),
                'average_margin_trend': self._calculate_trend(
                    sum(p['margin'] for p in current_period['by_product']) / len(current_period['by_product']) if current_period['by_product'] else 0,
                    sum(p['margin'] for p in previous_period['by_product']) / len(previous_period['by_product']) if previous_period and previous_period['by_product'] else 0
                )
            }
        }