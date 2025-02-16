from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db import transaction
from django.utils import timezone
from django.contrib import messages
from django.db.models import Q, F, Sum, Count, Func
from decimal import Decimal
import json
import decimal
from django.urls import reverse
from django.db.models.functions import Abs
from django.core.exceptions import ValidationError

from users.mixins import AdminRequiredMixin
from products.models import Product, Category
from .models import InventoryCount, InventoryCountDetail, InventoryAdjustment
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from products.models import ProductStock


class InventoryCountListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    """Lista todos los conteos de inventario con filtros y búsqueda."""
    model = InventoryCount
    template_name = 'inventory_management/inventory_list.html'
    context_object_name = 'inventories'
    paginate_by = 10

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search', '')
        status = self.request.GET.get('status', '')
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.select_related('created_by', 'reviewed_by')

class InventoryCountCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = InventoryCount
    template_name = 'inventory_management/inventory_form.html'
    fields = ['name', 'categories', 'notes']
    success_url = reverse_lazy('inventory_management:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Obtenemos todas las categorías sin filtrar por is_active
        context['categories'] = Category.objects.all().order_by('name')
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                # Guardar el conteo principal
                self.object = form.save(commit=False)
                self.object.created_by = self.request.user
                self.object.save()
                form.save_m2m()  # Guardar relaciones many-to-many

                # Obtener productos según las categorías seleccionadas
                from products.models import Product
                
                products_query = Product.objects.filter(is_active=True)
                if self.object.categories.exists():
                    products_query = products_query.filter(
                        category__in=self.object.categories.all()
                    )

                # Crear detalles para cada producto
                details = []
                for product in products_query:
                    expected_qty = product.bulk_stock if product.is_bulk else product.stock
                    details.append(
                        InventoryCountDetail(
                            inventory_count=self.object,
                            product=product,
                            expected_quantity=expected_qty
                        )
                    )
                
                if details:
                    InventoryCountDetail.objects.bulk_create(details)

                messages.success(
                    self.request,
                    'Conteo de inventario creado exitosamente.'
                )
                return super().form_valid(form)
                
        except Exception as e:
            messages.error(
                self.request,
                f'Error al crear el conteo de inventario: {str(e)}'
            )
            return self.form_invalid(form)

class InventoryCountDetailView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = InventoryCount
    template_name = 'inventory_management/inventory_detail.html'
    context_object_name = 'inventory'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory = self.object
        
        # Agregar conteos al contexto
        context.update({
            'total_products': inventory.details.count(),
            'counted_products': inventory.details.filter(actual_quantity__isnull=False).count(),
            'pending_products': inventory.details.filter(actual_quantity__isnull=True).count(),
            'products_with_differences': inventory.details.exclude(
                actual_quantity=F('expected_quantity')
            ).exclude(actual_quantity__isnull=True).count(),
        })
        
        return context

class InventoryCountProcessView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    """Vista para realizar el proceso de conteo."""
    model = InventoryCount
    template_name = 'inventory_management/inventory_count_process.html'
    context_object_name = 'inventory'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['uncounted_products'] = self.object.details.filter(
            actual_quantity__isnull=True
        ).count()
        return context

# class RequestRecountView(LoginRequiredMixin, AdminRequiredMixin, View):
#     def post(self, request, detail_id):
#         try:
#             print(f"Recibiendo solicitud de reconteo para detail_id: {detail_id}")
#             # Extraer razón del cuerpo de la solicitud
#             data = json.loads(request.body)
#             reason = data.get('reason', '')
#             print(f"Razón recibida: {reason}")
            
#             with transaction.atomic():
#                 detail = get_object_or_404(InventoryCountDetail, id=detail_id)
#                 print(f"Detalle encontrado: {detail}")
                
#                 # Actualizar detalle para reconteo
#                 detail.recount_requested = True
#                 detail.recount_reason = reason
#                 detail.recount_requested_by = request.user
#                 detail.actual_quantity = None  # Este es el cambio clave
#                 detail.counted_by = None
#                 detail.last_counted_at = None
#                 detail.notes = ''
#                 detail.save()
                
#                 print("Detalle actualizado exitosamente")

#                 return JsonResponse({
#                     'status': 'success',
#                     'message': 'Reconteo solicitado exitosamente'
#                 })

#         except Exception as e:
#             print(f"Error en solicitud de reconteo: {str(e)}")
#             return JsonResponse({
#                 'status': 'error',
#                 'message': str(e)
#             }, status=500)


class InventoryDashboardView(LoginRequiredMixin, AdminRequiredMixin, TemplateView):
    """Dashboard con estadísticas de inventarios."""
    template_name = 'inventory_management/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Estadísticas generales
        context['total_counts'] = InventoryCount.objects.count()
        context['in_progress_counts'] = InventoryCount.objects.filter(
            status='in_progress'
        ).count()
        
        # Últimos inventarios
        context['recent_counts'] = InventoryCount.objects.order_by(
            '-date_started'
        )[:5]
        
        # Productos con más diferencias
        context['problematic_products'] = InventoryCountDetail.objects.exclude(
            actual_quantity__isnull=True
        ).exclude(
            actual_quantity=F('expected_quantity')
        ).values(
            'product__name'
        ).annotate(
            difference_count=Count('id')
        ).order_by('-difference_count')[:10]
        
        return context
    
class SearchProductsView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        try:
            inventory_id = request.GET.get('inventory_id')
            query = request.GET.get('query', '').strip()

            if not inventory_id or not query:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Parámetros inválidos'
                }, status=400)

            inventory = get_object_or_404(InventoryCount, pk=inventory_id)
            
            # Modificar la consulta para incluir productos que:
            # 1. No han sido contados (actual_quantity is null)
            # 2. Están marcados para reconteo (recount_requested=True)
            details = inventory.details.select_related('product', 'product__category').filter(
                Q(actual_quantity__isnull=True) | 
                Q(recount_requested=True)
            ).filter(
                Q(product__name__icontains=query) |
                Q(product__brand__icontains=query)
            )[:10]

            results = []
            for detail in details:
                # Determinar la cantidad esperada según el tipo de producto
                expected_qty = (detail.product.bulk_stock 
                              if detail.product.is_bulk 
                              else detail.product.stock)
                
                results.append({
                    'id': detail.id,
                    'product_name': detail.product.name,
                    'product_code': detail.product.brand,
                    'category': detail.product.category.name,
                    'expected_quantity': float(expected_qty),
                    'is_bulk': detail.product.is_bulk,
                    'unit_type': 'kilos' if detail.product.is_bulk else 'unidades',
                    'recount_requested': detail.recount_requested,
                    'recount_count': detail.recount_count
                })

            return JsonResponse({
                'status': 'success',
                'results': results
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class SaveCountView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            detail_id = data.get('detail_id')
            quantity = data.get('quantity')
            notes = data.get('notes', '')

            with transaction.atomic():
                detail = InventoryCountDetail.objects.select_related('product').get(pk=detail_id)
                
                try:
                    if detail.product.is_bulk:
                        quantity = Decimal(str(quantity))
                    else:
                        quantity = Decimal(str(int(float(quantity))))
                except (ValueError, decimal.InvalidOperation):
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Cantidad inválida'
                    }, status=400)

                # Actualizar detalle
                detail.actual_quantity = quantity
                detail.notes = notes
                detail.counted_by = request.user
                detail.last_counted_at = timezone.now()
                detail.recount_requested = False  # Importante resetear el flag
                detail.save()

                # Actualizar estado del inventario si es necesario
                inventory = detail.inventory_count
                if inventory.status == 'draft':
                    inventory.status = 'in_progress'
                    inventory.save()

                # Calcular progreso actualizado
                total_products = inventory.details.count()
                counted_products = inventory.details.exclude(
                    actual_quantity__isnull=True
                ).exclude(
                    recount_requested=True  # No contar productos que necesitan reconteo
                ).count()
                
                # Preparar datos de respuesta
                progress_data = {
                    'percentage': round((counted_products / total_products) * 100),
                    'total': total_products,
                    'counted': counted_products,
                    'differences': inventory.details.filter(
                        ~Q(actual_quantity=F('expected_quantity')),
                        actual_quantity__isnull=False
                    ).count(),
                    'pending': total_products - counted_products
                }

                return JsonResponse({
                    'status': 'success',
                    'message': 'Conteo guardado exitosamente',
                    'progress': progress_data
                })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

class ApproveAdjustmentView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, pk):
        try:
            adjustment = InventoryAdjustment.objects.get(pk=pk)
            
            with transaction.atomic():
                adjustment.status = 'approved'
                adjustment.approved_by = request.user
                adjustment.save()
                
                # Actualizar stock del producto
                product = adjustment.product
                product.stock += adjustment.quantity_difference
                product.save()
                
            return JsonResponse({
                'status': 'success',
                'message': 'Ajuste aprobado exitosamente'
            })
            
        except InventoryAdjustment.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': 'Ajuste no encontrado'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al aprobar ajuste: {str(e)}'
            }, status=500)
        
class InventoryCountUpdateView(LoginRequiredMixin, AdminRequiredMixin, UpdateView):
    model = InventoryCount
    template_name = 'inventory_management/inventory_form.html'
    fields = ['name', 'categories', 'notes']
    success_url = reverse_lazy('inventory_management:list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all()
        return context

    def form_valid(self, form):
        try:
            with transaction.atomic():
                messages.success(self.request, 'Inventario actualizado exitosamente.')
                return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Error al actualizar el inventario: {str(e)}')
            return self.form_invalid(form)

    def get_queryset(self):
        # Solo permitir editar inventarios en estado 'draft'
        return super().get_queryset().filter(status='draft')
    
class InventoryCountDeleteView(LoginRequiredMixin, AdminRequiredMixin, DeleteView):
    model = InventoryCount
    template_name = 'inventory_management/inventory_confirm_delete.html'
    success_url = reverse_lazy('inventory_management:list')

    def get_queryset(self):
        # Solo permitir eliminar inventarios en estado 'draft'
        return super().get_queryset().filter(status='draft')

    def delete(self, request, *args, **kwargs):
        try:
            response = super().delete(request, *args, **kwargs)
            messages.success(request, 'Inventario eliminado exitosamente.')
            return response
        except Exception as e:
            messages.error(request, f'Error al eliminar el inventario: {str(e)}')
            return redirect('inventory_management:list')
        
class InventoryAdjustmentListView(LoginRequiredMixin, AdminRequiredMixin, ListView):
    model = InventoryAdjustment
    template_name = 'inventory_management/adjustment_list.html'
    context_object_name = 'adjustments'
    paginate_by = 10

    def get_queryset(self):
        queryset = InventoryAdjustment.objects.filter(
            inventory_count_id=self.kwargs['pk']
        ).select_related('product', 'created_by', 'approved_by')
        
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inventory'] = get_object_or_404(InventoryCount, pk=self.kwargs['pk'])
        context['status_choices'] = InventoryAdjustment.ADJUSTMENT_STATUS
        return context
    
class InventoryAdjustmentCreateView(LoginRequiredMixin, AdminRequiredMixin, CreateView):
    model = InventoryAdjustment
    template_name = 'inventory_management/adjustment_form.html'
    fields = ['product', 'quantity_difference', 'justification']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['inventory'] = get_object_or_404(InventoryCount, pk=self.kwargs['pk'])
        return context

    def form_valid(self, form):
        inventory = get_object_or_404(InventoryCount, pk=self.kwargs['pk'])
        form.instance.inventory_count = inventory
        form.instance.created_by = self.request.user
        
        try:
            # Calcular el monto del ajuste
            product = form.instance.product
            quantity = form.instance.quantity_difference
            price = product.weighted_average_price
            form.instance.amount = abs(quantity * price)
            
            messages.success(self.request, 'Ajuste creado exitosamente.')
            return super().form_valid(form)
        except Exception as e:
            messages.error(self.request, f'Error al crear el ajuste: {str(e)}')
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse_lazy('inventory_management:adjustments', kwargs={'pk': self.kwargs['pk']})

class InventoryReportView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = InventoryCount
    template_name = 'inventory_management/inventory_report.html'
    context_object_name = 'inventory'

    def get_template_names(self):
        if self.request.GET.get('format') == 'pdf':
            return ['inventory_management/print_report.html']
        return [self.template_name]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        inventory = self.object
        details = inventory.details.all()

        # Calcular estadísticas generales
        total_items = details.count()
        counted_items = details.exclude(actual_quantity__isnull=True).count()
        
        context.update({
            'total_items': total_items,
            'counted_items': counted_items,
            'completion_percentage': round((counted_items / total_items * 100) if total_items > 0 else 0, 2),
            
            # Agrupar diferencias por categoría
            'category_differences': self.get_category_differences(details),
            
            # Productos con mayores diferencias
            'top_differences': self.get_top_differences(details),
            
            # Valor total de las diferencias
            'total_difference_value': self.get_total_difference_value(details)
        })

        # if self.request.GET.get('format') == 'pdf':
        #     self.template_name = 'inventory_management/print_report.html'  
        
        return context
    
    def get_category_differences(self, details):
        differences_by_category = {}
        
        for detail in details.select_related('product__category'):
            if detail.actual_quantity is None:
                continue
                
            category_name = detail.product.category.name
            difference = detail.actual_quantity - detail.expected_quantity
            value_difference = difference * detail.product.purchase_price  # Usando precio de compra
            
            if category_name not in differences_by_category:
                differences_by_category[category_name] = {
                    'total_difference': 0,
                    'value_difference': 0
                }
            
            differences_by_category[category_name]['total_difference'] += difference
            differences_by_category[category_name]['value_difference'] += value_difference
        
        # Convertir a lista para la plantilla
        return [
            {
                'category': cat,
                'total_difference': data['total_difference'],
                'value_difference': data['value_difference']
            }
            for cat, data in differences_by_category.items()
        ]
    
    def get_top_differences(self, details):
        return details.exclude(
            actual_quantity__isnull=True
        ).annotate(
            difference=F('actual_quantity') - F('expected_quantity'),
            abs_difference=Func(
                F('actual_quantity') - F('expected_quantity'),
                function='ABS'
            )
        ).order_by('-abs_difference')[:10]
    
    def get_total_difference_value(self, details):
        total = 0
        for detail in details.exclude(actual_quantity__isnull=True):
            difference = detail.actual_quantity - detail.expected_quantity
            total += difference * detail.product.purchase_price
        return total
    
class StartCountView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, pk):
        try:
            inventory = get_object_or_404(InventoryCount, pk=pk)
            
            if inventory.status != 'draft':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Este inventario no está en estado borrador'
                }, status=400)
            
            with transaction.atomic():
                # Cambiar estado a en progreso
                inventory.status = 'in_progress'
                inventory.save()
                
                messages.success(request, 'Conteo iniciado exitosamente')
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Conteo iniciado exitosamente',
                    'redirect_url': reverse('inventory_management:count_process', args=[pk])
                })
        
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al iniciar el conteo: {str(e)}'
            }, status=500)
        
class PauseInventoryView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, pk):
        try:
            inventory = get_object_or_404(InventoryCount, pk=pk)
            
            if inventory.status != 'in_progress':
                return JsonResponse({
                    'status': 'error',
                    'message': 'Este inventario no está en proceso'
                }, status=400)
            
            # Guardar el estado actual
            inventory.save()
            
            return JsonResponse({
                'status': 'success',
                'message': 'Inventario pausado exitosamente',
                'redirect_url': reverse('inventory_management:detail', args=[pk])
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al pausar el inventario: {str(e)}'
            }, status=500)

class FinishInventoryView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, pk):
        try:
            with transaction.atomic():
                inventory = get_object_or_404(InventoryCount, pk=pk)
                
                if inventory.status != 'in_progress':
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Este inventario no está en proceso'
                    }, status=400)

                # Verificar si hay productos sin contar
                pending_products = inventory.details.filter(actual_quantity__isnull=True).exists()
                if pending_products:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'No se puede finalizar el inventario. Hay productos pendientes por contar.'
                    }, status=400)

                # Actualizar estado del inventario
                inventory.status = 'completed'
                inventory.date_finished = timezone.now()
                inventory.save()
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Inventario finalizado exitosamente',
                    'redirect_url': reverse('inventory_management:detail', args=[pk])
                })
                
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al finalizar el inventario: {str(e)}'
            }, status=500)
        
class DashboardDataView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        try:
            # Obtener parámetros de filtrado
            period = int(request.GET.get('period', 180))  # Default 6 meses
            trend_type = request.GET.get('trend_type', 'differences')
            category_metric = request.GET.get('category_metric', 'quantity')
            problem_metric = request.GET.get('problem_metric', 'difference')

            # Calcular fechas
            end_date = timezone.now()
            start_date = end_date - timedelta(days=period)

            # Obtener inventarios en el período
            inventories = InventoryCount.objects.filter(
                date_started__gte=start_date,
                date_started__lte=end_date
            )
            completed_inventories = inventories.filter(status='completed')

            # Obtener resumen de datos
            summary = self.get_summary_data(completed_inventories, period)

            data = {
                'summary': summary,  # Usar el summary que retorna get_summary_data
                'trends': self.get_trends_data(completed_inventories, trend_type),
                'categories': self.get_category_data(completed_inventories, category_metric),
                'problematic_products': self.get_problematic_products(completed_inventories, problem_metric)
            }

            print("Enviando datos:", data)  # Debug
            return JsonResponse(data)

        except Exception as e:
            print(f"Error en dashboard: {str(e)}")  # Debug
            return JsonResponse({
                'error': str(e),
                'summary': {
                    'totalInventories': {'value': 0, 'trend': 0},
                    'inProgressCount': {'value': 0, 'trend': 0},
                    'totalDifferences': {'value': 0, 'trend': 0},
                    'totalValue': {'value': 0, 'trend': 0}
                },
                'trends': {'labels': [], 'positive_differences': [], 'negative_differences': []},
                'categories': {'labels': [], 'values': []},
                'problematic_products': []
            }, status=500)

    def get_summary_data(self, inventories, period):
        """Calcula resumen con tendencias comparando con período anterior"""
        current_period = inventories
        previous_start = timezone.now() - timedelta(days=period*2)
        previous_end = timezone.now() - timedelta(days=period)
        previous_period = InventoryCount.objects.filter(
            date_started__gte=previous_start,
            date_started__lte=previous_end,
            status='completed'
        )

        def calculate_trend(current, previous):
            if previous == 0:
                return 0
            return ((current - previous) / previous) * 100

        # Total Inventarios
        current_total = current_period.count()
        previous_total = previous_period.count()
        
        # Diferencias y Valores
        current_differences = 0
        current_value = 0
        previous_differences = 0
        previous_value = 0

        for inventory in current_period:
            for detail in inventory.details.all():
                if detail.actual_quantity is not None:
                    diff = abs(detail.actual_quantity - detail.expected_quantity)
                    current_differences += diff
                    current_value += diff * detail.product.purchase_price

        for inventory in previous_period:
            for detail in inventory.details.all():
                if detail.actual_quantity is not None:
                    diff = abs(detail.actual_quantity - detail.expected_quantity)
                    previous_differences += diff
                    previous_value += diff * detail.product.purchase_price

        # Asegurar que no tengamos valores None
        current_value = current_value or 0
        previous_value = previous_value or 0
        current_differences = current_differences or 0
        previous_differences = previous_differences or 0

        return {
            'totalInventories': {
                'value': current_total,
                'trend': calculate_trend(current_total, previous_total)
            },
            'inProgressCount': {
                'value': inventories.filter(status='in_progress').count(),
                'trend': 0  # No hay tendencia para en progreso
            },
            'totalDifferences': {
                'value': current_differences,
                'trend': calculate_trend(current_differences, previous_differences)
            },
            'totalValue': {
                'value': round(current_value),  # Redondear valores monetarios
                'trend': round(calculate_trend(current_value, previous_value), 2)
            }
        }

    def get_trends_data(self, inventories, trend_type):
        labels = []
        positive_differences = []
        negative_differences = []

        # Agrupar por mes
        month_data = {}
        
        for inventory in inventories:
            month = inventory.date_started.strftime('%Y-%m')
            if month not in month_data:
                month_data[month] = {
                    'pos_diff': 0,
                    'neg_diff': 0,
                    'value_pos': 0,
                    'value_neg': 0,
                    'count': 0
                }

            month_data[month]['count'] += 1
            
            for detail in inventory.details.all():
                if detail.actual_quantity is not None:
                    diff = detail.actual_quantity - detail.expected_quantity
                    value = abs(diff * detail.product.purchase_price)
                    
                    if diff > 0:
                        month_data[month]['pos_diff'] += diff
                        month_data[month]['value_pos'] += value
                    elif diff < 0:
                        month_data[month]['neg_diff'] += abs(diff)
                        month_data[month]['value_neg'] += value

        # Ordenar meses
        sorted_months = sorted(month_data.keys())
        labels = sorted_months

        # Seleccionar datos según tipo
        if trend_type == 'differences':
            positive_differences = [month_data[m]['pos_diff'] for m in sorted_months]
            negative_differences = [month_data[m]['neg_diff'] for m in sorted_months]
        elif trend_type == 'value':
            positive_differences = [month_data[m]['value_pos'] for m in sorted_months]
            negative_differences = [month_data[m]['value_neg'] for m in sorted_months]
        else:  # counts
            positive_differences = [month_data[m]['count'] for m in sorted_months]
            negative_differences = [0 for _ in sorted_months]

        return {
            'labels': labels,
            'positive_differences': positive_differences,
            'negative_differences': negative_differences
        }

    def get_category_data(self, inventories, metric):
        category_data = {}
        
        for inventory in inventories:
            for detail in inventory.details.all():
                if detail.actual_quantity is None:
                    continue
                    
                category_name = detail.product.category.name
                if category_name not in category_data:
                    category_data[category_name] = {
                        'quantity': 0,
                        'value': 0,
                        'count': 0
                    }
                
                diff = detail.actual_quantity - detail.expected_quantity
                value = abs(diff * detail.product.purchase_price)
                
                category_data[category_name]['quantity'] += abs(diff)
                category_data[category_name]['value'] += value
                category_data[category_name]['count'] += 1

        # Calcular frecuencias
        total_count = sum(data['count'] for data in category_data.values())
        
        labels = list(category_data.keys())
        if metric == 'quantity':
            values = [data['quantity'] for data in category_data.values()]
        elif metric == 'value':
            values = [data['value'] for data in category_data.values()]
        else:  # frequency
            values = [(data['count']/total_count)*100 for data in category_data.values()]

        return {
            'labels': labels,
            'values': values
        }

    def get_problematic_products(self, inventories, metric):
        product_data = {}
        
        for inventory in inventories:
            for detail in inventory.details.all():
                if detail.actual_quantity is None:
                    continue
                    
                product = detail.product
                if product.id not in product_data:
                    product_data[product.id] = {
                        'name': product.name,
                        'code': product.brand,
                        'total_difference': 0,
                        'value_difference': 0,
                        'count': 0
                    }
                
                diff = detail.actual_quantity - detail.expected_quantity
                value = diff * product.purchase_price
                
                product_data[product.id]['total_difference'] += diff
                product_data[product.id]['value_difference'] += value
                product_data[product.id]['count'] += 1

        # Convertir a lista y calcular frecuencias
        products = list(product_data.values())
        total_count = sum(p['count'] for p in products)
        
        for product in products:
            product['frequency'] = round((product['count'] / total_count * 100), 2) if total_count > 0 else 0
            product['total_difference'] = round(float(product['total_difference']), 2)
            product['value_difference'] = round(float(product['value_difference']), 2)

        # Ordenar según métrica
        if metric == 'difference':
            key = lambda x: abs(x['total_difference'])
        elif metric == 'value':
            key = lambda x: abs(x['value_difference'])
        else:  # frequency
            key = lambda x: x['frequency']

        return sorted(products, key=key, reverse=True)[:10]
    
class ExportProblematicProductsView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        try:
            metric = request.GET.get('metric', 'difference')
            period = int(request.GET.get('period', 180))

            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos Problemáticos"

            # Estilos
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            header_font = Font(bold=True)

            # Encabezados
            headers = [
                'Producto',
                'Categoría',
                'Diferencia Total',
                'Valor Diferencia',
                'Frecuencia (%)',
                'Último Conteo',
                'Tendencia'
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Obtener datos
            end_date = timezone.now()
            start_date = end_date - timedelta(days=period)
            
            inventories = InventoryCount.objects.filter(
                date_started__gte=start_date,
                status='completed'
            )

            product_data = {}
            for inventory in inventories:
                for detail in inventory.details.select_related('product', 'product__category').all():
                    if detail.actual_quantity is None:
                        continue

                    product = detail.product
                    if product.id not in product_data:
                        product_data[product.id] = {
                            'name': product.name,
                            'category': product.category.name,
                            'total_difference': 0,
                            'value_difference': 0,
                            'count': 0,
                            'last_count': None
                        }

                    data = product_data[product.id]
                    diff = detail.actual_quantity - detail.expected_quantity
                    
                    data['total_difference'] += diff
                    data['value_difference'] += diff * product.purchase_price
                    data['count'] += 1
                    
                    if data['last_count'] is None or detail.last_counted_at > data['last_count']:
                        data['last_count'] = detail.last_counted_at

            # Calcular frecuencias y ordenar
            products = list(product_data.values())
            total_counts = sum(p['count'] for p in products)
            
            for product in products:
                product['frequency'] = (product['count'] / total_counts) * 100

            # Ordenar según métrica seleccionada
            if metric == 'difference':
                products.sort(key=lambda x: abs(x['total_difference']), reverse=True)
            elif metric == 'value':
                products.sort(key=lambda x: abs(x['value_difference']), reverse=True)
            else:  # frequency
                products.sort(key=lambda x: x['frequency'], reverse=True)

            # Llenar datos
            for row, product in enumerate(products, 2):
                ws.cell(row=row, column=1, value=product['name'])
                ws.cell(row=row, column=2, value=product['category'])
                ws.cell(row=row, column=3, value=float(product['total_difference']))
                ws.cell(row=row, column=4, value=float(product['value_difference']))
                ws.cell(row=row, column=5, value=round(product['frequency'], 2))
                ws.cell(row=row, column=6, value=product['last_count'].strftime('%d/%m/%Y %H:%M') if product['last_count'] else '-')
                
                # Calcular tendencia
                if product['total_difference'] > 0:
                    ws.cell(row=row, column=7, value="↑")
                elif product['total_difference'] < 0:
                    ws.cell(row=row, column=7, value="↓")
                else:
                    ws.cell(row=row, column=7, value="=")

            # Ajustar anchos de columna
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2

            # Preparar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=productos_problematicos.xlsx'
            
            wb.save(response)
            return response

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)



# Vista para edición masiva
class EditCompleteInventoryView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request, pk):
        inventory = get_object_or_404(InventoryCount, pk=pk)
        if inventory.status != 'completed':
            messages.error(request, 'Solo se pueden editar inventarios completados')
            return redirect('inventory_management:detail', pk=pk)
            
        context = {
            'inventory': inventory,
            'details': inventory.details.all().select_related('product', 'product__category')
        }
        return render(request, 'inventory_management/edit_inventory.html', context)

    def post(self, request, pk):
        try:
            data = json.loads(request.body)
            updates = data.get('updates', [])
            
            with transaction.atomic():
                inventory = get_object_or_404(InventoryCount, pk=pk)
                
                for update in updates:
                    if not update.get('modified'):
                        continue
                        
                    detail = get_object_or_404(
                        InventoryCountDetail,
                        id=update['detail_id'],
                        inventory_count=inventory
                    )
                    
                    # Guardar historial actual
                    previous_count = {
                        'quantity': str(detail.actual_quantity),
                        'counted_by': detail.counted_by.get_full_name() if detail.counted_by else None,
                        'date': detail.last_counted_at.isoformat() if detail.last_counted_at else None,
                        'reason': detail.notes
                    }
                    
                    history = detail.previous_counts or {}
                    history[str(len(history))] = previous_count
                    
                    # Convertir y validar la cantidad
                    try:
                        new_quantity = update['quantity']
                        if detail.product.is_bulk:
                            # Para productos a granel, permitir decimales
                            quantity = Decimal(str(float(new_quantity)))
                        else:
                            # Para productos por unidad, forzar enteros
                            quantity = Decimal(str(int(float(new_quantity))))
                    except (ValueError, decimal.InvalidOperation) as e:
                        return JsonResponse({
                            'status': 'error',
                            'message': f'Error en la cantidad para {detail.product.name}: {str(e)}'
                        }, status=400)
                    
                    # Actualizar con nuevos valores
                    detail.actual_quantity = quantity
                    detail.notes = update.get('reason', '')
                    detail.last_counted_at = timezone.now()
                    detail.counted_by = request.user
                    detail.previous_counts = history
                    detail.save()
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Cambios guardados exitosamente'
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Error en el formato de los datos'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Error al guardar los cambios: {str(e)}'
            }, status=500)

# Vista para edición individual
class EditInventoryDetailView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request, detail_id):
        try:
            detail = get_object_or_404(InventoryCountDetail, id=detail_id)
            data = json.loads(request.body)
            
            with transaction.atomic():
                # Guardar historial
                previous_count = {
                    'quantity': str(detail.actual_quantity),
                    'counted_by': detail.counted_by.get_full_name() if detail.counted_by else None,
                    'date': detail.last_counted_at.isoformat() if detail.last_counted_at else None,
                    'reason': detail.notes
                }
                
                history = detail.previous_counts or {}
                history[str(len(history))] = previous_count
                
                # Actualizar
                detail.actual_quantity = Decimal(str(data['quantity']))
                detail.notes = data.get('reason', '')
                detail.last_counted_at = timezone.now()
                detail.counted_by = request.user
                detail.previous_counts = history
                detail.save()
                
                return JsonResponse({
                    'status': 'success',
                    'message': 'Cambios guardados exitosamente'
                })
                
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)
        

class ProcessInventoryAdjustmentsView(LoginRequiredMixin, AdminRequiredMixin, DetailView):
    model = InventoryCount
    template_name = 'inventory_management/process_adjustments.html'
    context_object_name = 'inventory'

    def get_queryset(self):
        # Solo permitir procesar ajustes de inventarios completados
        return super().get_queryset().filter(status='completed')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        details = self.object.details.exclude(
            actual_quantity=F('expected_quantity')
        ).exclude(actual_quantity__isnull=True)
        
        # Agregar información de ajustes previos
        for detail in details:
            detail.is_adjusted = InventoryAdjustment.objects.filter(
                inventory_count_detail=detail,
                applied=True
            ).exists()
        
        context['details'] = details
        context['total_differences'] = details.count()
        return context

class ApplyAdjustmentsView(LoginRequiredMixin, AdminRequiredMixin, View):
    def post(self, request):
        try:
            data = json.loads(request.body)
            detail_ids = data.get('detail_ids', [])
            
            with transaction.atomic():
                details = InventoryCountDetail.objects.filter(
                    id__in=detail_ids,
                    inventory_count__status='completed'
                ).select_related('product', 'inventory_count')

                for detail in details:
                    # Verificar si el producto ya fue ajustado
                    if InventoryAdjustment.objects.filter(
                        inventory_count_detail=detail,
                        applied=True
                    ).exists():
                        raise ValidationError(
                            f'El producto {detail.product.name} ya fue ajustado anteriormente'
                        )

                    if detail.actual_quantity is None:
                        continue

                    difference = detail.actual_quantity - detail.expected_quantity
                    
                    # Generar un batch number único para cada ajuste
                    batch_number = f"INV-{detail.inventory_count.id}-{detail.id}-{timezone.now().strftime('%Y%m%d%H%M%S')}"
                    
                    # Crear el ajuste
                    adjustment = InventoryAdjustment.objects.create(
                        inventory_count=detail.inventory_count,
                        product=detail.product,
                        quantity_difference=difference,
                        amount=abs(difference * detail.product.purchase_price),
                        created_by=request.user,
                        status='approved',
                        justification=f'Ajuste automático por inventario {detail.inventory_count.name}',
                        inventory_count_detail=detail,
                        applied=True
                    )

                    # Actualizar stock del producto
                    product = detail.product
                    if product.is_bulk:
                        product.bulk_stock = detail.actual_quantity
                    else:
                        product.stock = detail.actual_quantity
                    product.save()

                    # Crear movimiento en ProductStock
                    changes = {
                        'Stock': {
                            'old': str(detail.expected_quantity),
                            'new': str(detail.actual_quantity)
                        }
                    }
                    
                    ProductStock.objects.create(
                        product=product,
                        movement_type='ADJ',
                        quantity=abs(difference),
                        remaining_quantity=0,
                        purchase_price=product.purchase_price,
                        batch_number=batch_number,
                        notes=f'Ajuste por inventario {detail.inventory_count.name}',
                        created_by=request.user,
                        date=timezone.now(),
                        changes_detail=changes
                    )

                return JsonResponse({
                    'status': 'success',
                    'message': 'Ajustes aplicados correctamente'
                })

        except ValidationError as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)